import os
import requests
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image


# Define the folder path where you want to save the Excel files
output_folder = '/Users/ysnk/Desktop/sem-8/bioproj/excel files'

# Ensure the output folder exists
os.makedirs(output_folder, exist_ok=True)

# Define the KEGG API base URL
base_url = 'http://rest.kegg.jp'

def get_pathway_map_diagram(pathway_id):
    # Construct the URL for the pathway map diagram
    diagram_url = f'{base_url}/get/{pathway_id}/image'
    
    # Send a GET request to the diagram URL
    response = requests.get(diagram_url)
    
    # Check if the request was successful
    if response.status_code == 200:
        # Return the image content as bytes
        return response.content
    else:
        # Print an error message if the request failed
        print(f'Error: Unable to retrieve pathway map diagram for {pathway_id}')
        return None


def write_pathway_map_diagrams_to_excel(pathway_map_diagrams):
    # Create a new Excel workbook
    wb = Workbook()
    
    # Iterate over each pathway and its corresponding map diagram
    for pathway_name, map_diagram in pathway_map_diagrams.items():
        # Create a new worksheet for the pathway
        ws = wb.create_sheet(title=pathway_name)
        
        # Convert the map diagram bytes to an image
        img = Image(BytesIO(map_diagram))
        
        # Add the image to the worksheet
        ws.add_image(img, 'A1')
    
    # Save the Excel file to the specified folder
    file_path = os.path.join(output_folder, 'pathway_map_diagrams.xlsx')
    wb.save(file_path)
    print(f'Pathway map diagrams for all pathways have been saved to {file_path}')


    # Create a new Excel workbook
    wb = Workbook()
    
    # Iterate over each pathway and its corresponding gene list
    for pathway_name, gene_list in pathway_gene_lists.items():
        # Create a new worksheet for the pathway
        ws = wb.create_sheet(title=pathway_name)
        
        # Write the headers for gene ID and gene name
        ws.cell(row=1, column=1, value='Gene ID')
        ws.cell(row=1, column=2, value='Gene Name')
        
        # Write the gene list to the worksheet
        for i, (gene_id, gene_name) in enumerate(gene_list, start=2):
            ws.cell(row=i, column=1, value=gene_id)
            ws.cell(row=i, column=2, value=gene_name)
    
    # Save the Excel file to the specified folder
    file_path = os.path.join(output_folder, 'gene_information.xlsx')
    wb.save(file_path)
    print(f'Gene information for all pathways has been saved to {file_path}')

# Pathway IDs and their corresponding names
pathway_data = {
    'mtu00010': 'Glycolysis or Gluconeogenesis',
    'mtu00020': 'Citrate cycle (TCA cycle)',
    'mtu00030': 'Pentose phosphate pathway',
    'mtu00040': 'Pentose and glucuronate interconversions',
    'mtu00051': 'Fructose and mannose metabolism',
    'mtu00052': 'Galactose metabolism',
    'mtu00053': 'Ascorbate and aldarate metabolism',
    'mtu00500': 'Starch and sucrose metabolism',
    'mtu00520': 'Amino sugar and nucleotide sugar metabolism',
    'mtu00620': 'Pyruvate metabolism',
    'mtu00630': 'Glyoxylate and dicarboxylate metabolism',
    'mtu00640': 'Propanoate metabolism',
    'mtu00650': 'Butanoate metabolism',
    'mtu00660': 'C5-Branched dibasic acid metabolism',
    'mtu00562': 'Inositol phosphate metabolism',
    'mtu00190': 'Oxidative phosphorylation',
    'mtu00680': 'Methane metabolism',
    'mtu00910': 'Nitrogen metabolism',
    'mtu00920': 'Sulfur metabolism',
    'mtu00061': 'Fatty acid biosynthesis',
    'mtu00071': 'Fatty acid degradation',
    'mtu00100': 'Steroid biosynthesis',
    'mtu00121': 'Secondary bile acid biosynthesis',
    'mtu00561': 'Glycerolipid metabolism',
    'mtu00564': 'Glycerophospholipid metabolism',
    'mtu00565': 'Ether lipid metabolism',
    'mtu00600': 'Sphingolipid metabolism',
    'mtu00592': 'alpha-Linolenic acid metabolism',
    'mtu01040': 'Biosynthesis of unsaturated fatty acids',
    'mtu00230': 'Purine metabolism',
    'mtu00240': 'Pyrimidine metabolism',
    'mtu00250': 'Alanine, aspartate and glutamate metabolism',
    'mtu00260': 'Glycine, serine and threonine metabolism',
    'mtu00270': 'Cysteine and methionine metabolism',
    'mtu00280': 'Valine, leucine and isoleucine degradation',
    'mtu00290': 'Valine, leucine and isoleucine biosynthesis',
    'mtu00300': 'Lysine biosynthesis',
    'mtu00310': 'Lysine degradation',
    'mtu00220': 'Arginine biosynthesis',
    'mtu00330': 'Arginine and proline metabolism',
    'mtu00340': 'Histidine metabolism',
    'mtu00350': 'Tyrosine metabolism',
    'mtu00360': 'Phenylalanine metabolism',
    'mtu00380': 'Tryptophan metabolism',
    'mtu00400': 'Phenylalanine, tyrosine and tryptophan biosynthesis',
    'mtu00410': 'beta-Alanine metabolism',
    'mtu00430': 'Taurine and hypotaurine metabolism',
    'mtu00450': 'Selenocompound metabolism',
    'mtu00460': 'Cyanoamino acid metabolism',
    'mtu00470': 'D-Amino acid metabolism',
    'mtu00480': 'Glutathione metabolism',
    'mtu00515': 'Mannose type O-glycan biosynthesis',
    'mtu00514': 'Other types of O-glycan biosynthesis',
    'mtu00540': 'Lipopolysaccharide biosynthesis',
    'mtu00542': 'O-Antigen repeat unit biosynthesis',
    'mtu00541': 'O-Antigen nucleotide sugar biosynthesis',
    'mtu00550': 'Peptidoglycan biosynthesis',
    'mtu00552': 'Teichoic acid biosynthesis',
    'mtu00571': 'Lipoarabinomannan (LAM) biosynthesis',
    'mtu00572': 'Arabinogalactan biosynthesis Mycobacterium',
    'mtu00543': 'Exopolysaccharide biosynthesis',
    'mtu00730': 'Thiamine metabolism',
    'mtu00740': 'Riboflavin metabolism',
    'mtu00750': 'Vitamin B6 metabolism',
    'mtu00760': 'Nicotinate and nicotinamide metabolism',
    'mtu00770': 'Pantothenate and CoA biosynthesis',
    'mtu00780': 'Biotin metabolism',
    'mtu00785': 'Lipoic acid metabolism',
    'mtu00790': 'Folate biosynthesis',
    'mtu00670': 'One carbon pool by folate',
    'mtu00860': 'Porphyrin metabolism',
    'mtu00130': 'Ubiquinone and other terpenoid quinone biosynthesis',
    'mtu00900': 'Terpenoid backbone biosynthesis',
    'mtu00906': 'Carotenoid biosynthesis',
    'mtu00903': 'Limonene degradation',
    'mtu00907': 'Pinene, camphor and geraniol degradation',
    'mtu00523': 'Polyketide sugar unit biosynthesis',
    'mtu01053': 'Biosynthesis of siderophore group nonribosomal peptides',
    'mtu00946': 'Degradation of flavonoids',
    'mtu00311': 'Penicillin and cephalosporin biosynthesis',
    'mtu00332': 'Carbapenem biosynthesis',
    'mtu00261': 'Monobactam biosynthesis',
    'mtu00521': 'Streptomycin biosynthesis',
    'mtu00525': 'Acarbose and validamycin biosynthesis',
    'mtu00401': 'Novobiocin biosynthesis',
    'mtu00999': 'Biosynthesis of various plant secondary metabolites',
    'mtu00362': 'Benzoate degradation',
    'mtu00627': 'Aminobenzoate degradation',
    'mtu00364': 'Fluorobenzoate degradation',
    'mtu00625': 'Chloroalkane and chloroalkene degradation',
    'mtu00361': 'Chlorocyclohexane and chlorobenzene degradation',
    'mtu00623': 'Toluene degradation',
    'mtu00622': 'Xylene degradation',
    'mtu00643': 'Styrene degradation',
    'mtu00930': 'Caprolactam degradation',
    'mtu00621': 'Dioxin degradation',
    'mtu00626': 'Naphthalene degradation',
    'mtu00984': 'Steroid degradation',
    'mtu00983': 'Drug metabolism - other enzymes',
    'mtu03020': 'RNA polymerase',
    'mtu03010': 'Ribosome',
    'mtu00970': 'Aminoacyl-tRNA biosynthesis',
    'mtu03060': 'Protein export',
    'mtu04122': 'Sulfur relay system',
    'mtu03050': 'Proteasome',
    'mtu03018': 'RNA degradation',
    'mtu03030': 'DNA replication',
    'mtu03410': 'Base excision repair',
    'mtu03420': 'Nucleotide excision repair',
    'mtu03430': 'Mismatch repair',
    'mtu03440': 'Homologous recombination',
    'mtu03450': 'Non-homologous end-joining',
    'mtu03250': 'Viral life cycle - HIV-1',
    'mtu02010': 'ABC transporters',
    'mtu03070': 'Bacterial secretion system',
    'mtu02020': 'Two-component system',
    'mtu02024': 'Quorum sensing',
    'mtu05152': 'Tuberculosis',
    'mtu01501': 'beta-Lactam resistance',
    'mtu01502': 'Vancomycin resistance',
    'mtu01503': 'Cationic antimicrobial peptide (CAMP) resistance',
    # Add more pathway IDs and names as needed
}

# Retrieve pathway map diagrams for each pathway and store in a dictionary
pathway_map_diagrams = {}
for pathway_id, pathway_name in pathway_data.items():
    map_diagram = get_pathway_map_diagram(pathway_id)
    if map_diagram:
        pathway_map_diagrams[pathway_name] = map_diagram

# Write all pathway map diagrams to the same Excel file with each diagram in a separate worksheet
write_pathway_map_diagrams_to_excel(pathway_map_diagrams)
